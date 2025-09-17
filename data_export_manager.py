#!/usr/bin/env python3
"""
Data Export and Management Module for DVAGO.pk Scraper
=====================================================

This module provides comprehensive data export functionality including:
- Multiple export formats (CSV, JSON, Excel, XML)
- Database management and optimization
- Data cleaning and validation
- Report generation
- Data analytics and insights

Author: GitHub Copilot
Date: September 17, 2025
"""

import pandas as pd
import json
import csv
import sqlite3
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
import os
import logging
from datetime import datetime
import numpy as np
from collections import Counter
import matplotlib.pyplot as plt
import seaborn as sns


class DataExportManager:
    """
    Comprehensive data export and management system
    """
    
    def __init__(self, database_path, output_dir):
        """
        Initialize the data export manager
        
        Args:
            database_path (str): Path to SQLite database
            output_dir (str): Output directory for exports
        """
        self.database_path = database_path
        self.output_dir = output_dir
        self.logger = logging.getLogger(__name__)
        
        # Ensure output directory exists
        os.makedirs(output_dir, exist_ok=True)
        
        # Create subdirectories for different export types
        self.create_export_directories()
        
        # Connect to database
        self.conn = sqlite3.connect(database_path)
        self.conn.row_factory = sqlite3.Row  # Enable column access by name
        
    def create_export_directories(self):
        """Create subdirectories for organized exports"""
        subdirs = [
            'csv_exports',
            'json_exports',
            'excel_exports',
            'xml_exports',
            'reports',
            'analytics',
            'images'
        ]
        
        for subdir in subdirs:
            os.makedirs(os.path.join(self.output_dir, subdir), exist_ok=True)
    
    def clean_and_validate_data(self):
        """Clean and validate data in the database"""
        self.logger.info("Starting data cleaning and validation...")
        
        cursor = self.conn.cursor()
        
        # Clean product data
        self.logger.info("Cleaning product data...")
        
        # Remove duplicate products
        cursor.execute('''
            DELETE FROM products 
            WHERE id NOT IN (
                SELECT MIN(id) 
                FROM products 
                GROUP BY url
            )
        ''')
        
        # Clean price data (remove negative prices, extreme outliers)
        cursor.execute('''
            UPDATE products 
            SET price_current = NULL 
            WHERE price_current <= 0 OR price_current > 1000000
        ''')
        
        cursor.execute('''
            UPDATE products 
            SET price_original = NULL 
            WHERE price_original <= 0 OR price_original > 1000000
        ''')
        
        # Fix discount percentages
        cursor.execute('''
            UPDATE products 
            SET discount_percentage = 
                CASE 
                    WHEN price_original > 0 AND price_current > 0 AND price_original > price_current
                    THEN ROUND(((price_original - price_current) * 100.0 / price_original), 2)
                    ELSE NULL
                END
            WHERE price_original IS NOT NULL AND price_current IS NOT NULL
        ''')
        
        # Clean category data
        self.logger.info("Cleaning category data...")
        
        # Remove duplicate categories
        cursor.execute('''
            DELETE FROM categories 
            WHERE id NOT IN (
                SELECT MIN(id) 
                FROM categories 
                GROUP BY url
            )
        ''')
        
        self.conn.commit()
        self.logger.info("Data cleaning completed")
    
    def export_to_csv(self, tables=None):
        """
        Export database tables to CSV files
        
        Args:
            tables (list): List of table names to export. If None, exports all tables.
        """
        self.logger.info("Exporting data to CSV format...")
        
        if tables is None:
            tables = ['categories', 'products', 'brands', 'product_images']
        
        csv_dir = os.path.join(self.output_dir, 'csv_exports')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        for table in tables:
            try:
                # Read data from database
                df = pd.read_sql_query(f"SELECT * FROM {table}", self.conn)
                
                if df.empty:
                    self.logger.warning(f"No data found in table: {table}")
                    continue
                
                # Clean data for CSV export
                df = self.clean_dataframe_for_export(df)
                
                # Export to CSV
                csv_file = os.path.join(csv_dir, f"{table}_{timestamp}.csv")
                df.to_csv(csv_file, index=False, encoding='utf-8')
                
                self.logger.info(f"Exported {len(df)} records from {table} to {csv_file}")
                
            except Exception as e:
                self.logger.error(f"Error exporting {table} to CSV: {str(e)}")
    
    def export_to_json(self, tables=None, pretty_print=True):
        """
        Export database tables to JSON files
        
        Args:
            tables (list): List of table names to export
            pretty_print (bool): Whether to format JSON with indentation
        """
        self.logger.info("Exporting data to JSON format...")
        
        if tables is None:
            tables = ['categories', 'products', 'brands']
        
        json_dir = os.path.join(self.output_dir, 'json_exports')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        for table in tables:
            try:
                cursor = self.conn.cursor()
                cursor.execute(f"SELECT * FROM {table}")
                
                # Convert to list of dictionaries
                data = [dict(row) for row in cursor.fetchall()]
                
                if not data:
                    self.logger.warning(f"No data found in table: {table}")
                    continue
                
                # Clean data for JSON export
                data = self.clean_data_for_json(data)
                
                # Export to JSON
                json_file = os.path.join(json_dir, f"{table}_{timestamp}.json")
                
                with open(json_file, 'w', encoding='utf-8') as f:
                    if pretty_print:
                        json.dump(data, f, indent=2, ensure_ascii=False, default=str)
                    else:
                        json.dump(data, f, ensure_ascii=False, default=str)
                
                self.logger.info(f"Exported {len(data)} records from {table} to {json_file}")
                
            except Exception as e:
                self.logger.error(f"Error exporting {table} to JSON: {str(e)}")
    
    def export_to_excel(self, filename=None):
        """
        Export all data to a comprehensive Excel file with multiple sheets
        
        Args:
            filename (str): Output filename. If None, generates timestamp-based name.
        """
        self.logger.info("Exporting data to Excel format...")
        
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"dvago_complete_data_{timestamp}.xlsx"
        
        excel_dir = os.path.join(self.output_dir, 'excel_exports')
        excel_file = os.path.join(excel_dir, filename)
        
        try:
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                # Export main tables
                tables = ['categories', 'products', 'brands', 'product_images']
                
                for table in tables:
                    df = pd.read_sql_query(f"SELECT * FROM {table}", self.conn)
                    
                    if not df.empty:
                        df = self.clean_dataframe_for_export(df)
                        df.to_excel(writer, sheet_name=table.title(), index=False)
                        self.logger.info(f"Added {len(df)} records from {table} to Excel")
                
                # Add summary sheet
                self.create_summary_sheet(writer)
                
                # Add analytics sheet
                self.create_analytics_sheet(writer)
            
            # Format the Excel file
            self.format_excel_file(excel_file)
            
            self.logger.info(f"Excel export completed: {excel_file}")
            
        except Exception as e:
            self.logger.error(f"Error creating Excel export: {str(e)}")
    
    def export_to_xml(self, tables=None):
        """
        Export database tables to XML files
        
        Args:
            tables (list): List of table names to export
        """
        self.logger.info("Exporting data to XML format...")
        
        if tables is None:
            tables = ['categories', 'products', 'brands']
        
        xml_dir = os.path.join(self.output_dir, 'xml_exports')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        for table in tables:
            try:
                cursor = self.conn.cursor()
                cursor.execute(f"SELECT * FROM {table}")
                
                # Create XML structure
                root = ET.Element(f"{table}_data")
                root.set('exported_at', datetime.now().isoformat())
                
                for row in cursor.fetchall():
                    item = ET.SubElement(root, table[:-1])  # Remove 's' from table name
                    
                    for key, value in dict(row).items():
                        if value is not None:
                            elem = ET.SubElement(item, key)
                            elem.text = str(value)
                
                # Write XML file
                xml_file = os.path.join(xml_dir, f"{table}_{timestamp}.xml")
                tree = ET.ElementTree(root)
                tree.write(xml_file, encoding='utf-8', xml_declaration=True)
                
                self.logger.info(f"Exported {table} to XML: {xml_file}")
                
            except Exception as e:
                self.logger.error(f"Error exporting {table} to XML: {str(e)}")
    
    def clean_dataframe_for_export(self, df):
        """Clean DataFrame for export"""
        # Handle NaN values
        df = df.fillna('')
        
        # Clean text columns
        text_columns = df.select_dtypes(include=['object']).columns
        for col in text_columns:
            df[col] = df[col].astype(str).str.strip()
            # Remove newlines and extra spaces
            df[col] = df[col].str.replace(r'\s+', ' ', regex=True)
        
        return df
    
    def clean_data_for_json(self, data):
        """Clean data for JSON export"""
        cleaned_data = []
        
        for item in data:
            cleaned_item = {}
            for key, value in item.items():
                if value is not None:
                    if isinstance(value, str):
                        # Clean text
                        cleaned_item[key] = value.strip()
                    else:
                        cleaned_item[key] = value
            cleaned_data.append(cleaned_item)
        
        return cleaned_data
    
    def create_summary_sheet(self, writer):
        """Create a summary sheet with key statistics"""
        summary_data = []
        
        # Get table counts
        tables = ['categories', 'products', 'brands']
        for table in tables:
            cursor = self.conn.cursor()
            cursor.execute(f"SELECT COUNT(*) as count FROM {table}")
            count = cursor.fetchone()[0]
            summary_data.append({'Table': table.title(), 'Record Count': count})
        
        # Get product statistics
        cursor = self.conn.cursor()
        
        # Price statistics
        cursor.execute('''
            SELECT 
                COUNT(*) as total_products,
                COUNT(price_current) as products_with_price,
                AVG(price_current) as avg_price,
                MIN(price_current) as min_price,
                MAX(price_current) as max_price
            FROM products 
            WHERE price_current > 0
        ''')
        
        price_stats = cursor.fetchone()
        if price_stats:
            summary_data.extend([
                {'Table': 'Product Statistics', 'Record Count': ''},
                {'Table': 'Total Products', 'Record Count': price_stats[0]},
                {'Table': 'Products with Price', 'Record Count': price_stats[1]},
                {'Table': 'Average Price (Rs.)', 'Record Count': f"{price_stats[2]:.2f}" if price_stats[2] else 'N/A'},
                {'Table': 'Min Price (Rs.)', 'Record Count': f"{price_stats[3]:.2f}" if price_stats[3] else 'N/A'},
                {'Table': 'Max Price (Rs.)', 'Record Count': f"{price_stats[4]:.2f}" if price_stats[4] else 'N/A'}
            ])
        
        # Create DataFrame and write to Excel
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
    
    def create_analytics_sheet(self, writer):
        """Create an analytics sheet with insights"""
        analytics_data = []
        
        cursor = self.conn.cursor()
        
        # Top categories by product count
        cursor.execute('''
            SELECT c.name, COUNT(p.id) as product_count
            FROM categories c
            LEFT JOIN products p ON c.id = p.category_id
            GROUP BY c.id, c.name
            ORDER BY product_count DESC
            LIMIT 10
        ''')
        
        top_categories = cursor.fetchall()
        
        analytics_data.append({'Metric': 'Top Categories by Product Count', 'Value': ''})
        for cat in top_categories:
            analytics_data.append({'Metric': cat[0], 'Value': cat[1]})
        
        analytics_data.append({'Metric': '', 'Value': ''})
        
        # Price range distribution
        cursor.execute('''
            SELECT 
                CASE 
                    WHEN price_current < 100 THEN 'Under Rs. 100'
                    WHEN price_current < 500 THEN 'Rs. 100-500'
                    WHEN price_current < 1000 THEN 'Rs. 500-1000'
                    WHEN price_current < 5000 THEN 'Rs. 1000-5000'
                    ELSE 'Over Rs. 5000'
                END as price_range,
                COUNT(*) as count
            FROM products 
            WHERE price_current > 0
            GROUP BY price_range
            ORDER BY MIN(price_current)
        ''')
        
        price_ranges = cursor.fetchall()
        
        analytics_data.append({'Metric': 'Price Range Distribution', 'Value': ''})
        for range_data in price_ranges:
            analytics_data.append({'Metric': range_data[0], 'Value': range_data[1]})
        
        # Create DataFrame and write to Excel
        analytics_df = pd.DataFrame(analytics_data)
        analytics_df.to_excel(writer, sheet_name='Analytics', index=False)
    
    def format_excel_file(self, excel_file):
        """Apply formatting to Excel file"""
        try:
            workbook = openpyxl.load_workbook(excel_file)
            
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                # Format headers
                header_font = Font(bold=True, color='FFFFFF')
                header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            workbook.save(excel_file)
            
        except Exception as e:
            self.logger.error(f"Error formatting Excel file: {str(e)}")
    
    def generate_data_report(self):
        """Generate a comprehensive data report"""
        self.logger.info("Generating comprehensive data report...")
        
        report_dir = os.path.join(self.output_dir, 'reports')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        report_file = os.path.join(report_dir, f"data_report_{timestamp}.html")
        
        # Generate report content
        report_html = self.create_html_report()
        
        # Write report to file
        with open(report_file, 'w', encoding='utf-8') as f:
            f.write(report_html)
        
        self.logger.info(f"Data report generated: {report_file}")
        return report_file
    
    def create_html_report(self):
        """Create HTML report content"""
        cursor = self.conn.cursor()
        
        # Get basic statistics
        stats = self.get_database_statistics()
        
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>DVAGO.pk Scraping Report</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                .header {{ background-color: #366092; color: white; padding: 20px; text-align: center; }}
                .section {{ margin: 20px 0; padding: 15px; border: 1px solid #ddd; }}
                .stats-table {{ width: 100%; border-collapse: collapse; }}
                .stats-table th, .stats-table td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                .stats-table th {{ background-color: #f2f2f2; }}
                .highlight {{ background-color: #e7f3ff; padding: 10px; border-left: 4px solid #366092; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>DVAGO.pk Scraping Report</h1>
                <p>Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            </div>
            
            <div class="section">
                <h2>Database Overview</h2>
                <table class="stats-table">
                    <tr><th>Metric</th><th>Value</th></tr>
                    <tr><td>Total Categories</td><td>{stats['total_categories']}</td></tr>
                    <tr><td>Total Products</td><td>{stats['total_products']}</td></tr>
                    <tr><td>Total Brands</td><td>{stats['total_brands']}</td></tr>
                    <tr><td>Products with Prices</td><td>{stats['products_with_price']}</td></tr>
                    <tr><td>Average Product Price</td><td>Rs. {stats['avg_price']:.2f}</td></tr>
                </table>
            </div>
            
            <div class="section">
                <h2>Data Quality</h2>
                <div class="highlight">
                    <p><strong>Completeness:</strong> {stats['completeness']:.1f}% of products have complete information</p>
                    <p><strong>Price Coverage:</strong> {stats['price_coverage']:.1f}% of products have pricing information</p>
                </div>
            </div>
            
            <div class="section">
                <h2>Export Files Generated</h2>
                <ul>
                    <li>CSV exports in: csv_exports/</li>
                    <li>JSON exports in: json_exports/</li>
                    <li>Excel exports in: excel_exports/</li>
                    <li>XML exports in: xml_exports/</li>
                </ul>
            </div>
        </body>
        </html>
        """
        
        return html_content
    
    def get_database_statistics(self):
        """Get comprehensive database statistics"""
        cursor = self.conn.cursor()
        
        stats = {}
        
        # Basic counts
        cursor.execute("SELECT COUNT(*) FROM categories")
        stats['total_categories'] = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM products")
        stats['total_products'] = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM brands")
        stats['total_brands'] = cursor.fetchone()[0]
        
        # Product statistics
        cursor.execute("SELECT COUNT(*) FROM products WHERE price_current IS NOT NULL AND price_current > 0")
        stats['products_with_price'] = cursor.fetchone()[0]
        
        cursor.execute("SELECT AVG(price_current) FROM products WHERE price_current > 0")
        avg_price = cursor.fetchone()[0]
        stats['avg_price'] = avg_price if avg_price else 0
        
        # Data quality metrics
        cursor.execute('''
            SELECT COUNT(*) FROM products 
            WHERE name IS NOT NULL AND name != '' 
            AND (price_current IS NOT NULL OR price_original IS NOT NULL)
        ''')
        complete_products = cursor.fetchone()[0]
        
        stats['completeness'] = (complete_products / stats['total_products'] * 100) if stats['total_products'] > 0 else 0
        stats['price_coverage'] = (stats['products_with_price'] / stats['total_products'] * 100) if stats['total_products'] > 0 else 0
        
        return stats
    
    def export_all_formats(self):
        """Export data in all available formats"""
        self.logger.info("Starting comprehensive data export in all formats...")
        
        try:
            # Clean data first
            self.clean_and_validate_data()
            
            # Export to all formats
            self.export_to_csv()
            self.export_to_json()
            self.export_to_excel()
            self.export_to_xml()
            
            # Generate report
            report_file = self.generate_data_report()
            
            self.logger.info("All exports completed successfully!")
            
            return {
                'csv_dir': os.path.join(self.output_dir, 'csv_exports'),
                'json_dir': os.path.join(self.output_dir, 'json_exports'),
                'excel_dir': os.path.join(self.output_dir, 'excel_exports'),
                'xml_dir': os.path.join(self.output_dir, 'xml_exports'),
                'report_file': report_file
            }
            
        except Exception as e:
            self.logger.error(f"Error during comprehensive export: {str(e)}")
            raise
    
    def __del__(self):
        """Cleanup database connection"""
        if hasattr(self, 'conn') and self.conn:
            self.conn.close()


def main():
    """Test the data export functionality"""
    import logging
    
    # Setup logging
    logging.basicConfig(level=logging.INFO)
    
    # Initialize export manager
    export_manager = DataExportManager(
        database_path="dvago_scraped_data/dvago_data.db",
        output_dir="dvago_scraped_data"
    )
    
    # Export all data
    export_paths = export_manager.export_all_formats()
    
    print("Export completed!")
    print("Generated files:")
    for export_type, path in export_paths.items():
        print(f"- {export_type}: {path}")


if __name__ == "__main__":
    main()